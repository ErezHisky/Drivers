import openpyxl
import time
from datetime import datetime

class MyExcel():
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.sheetNames = self.wb.sheetnames
        self.my_sheet = self.wb[self.sheetNames[0]]

    def get_sheet_names(self):
        return self.wb.sheetnames

    def openSheet(self, sheetName):
        self.my_sheet = self.wb[sheetName]

    def enter_data_to_raw(self):
        pass

    def create_new_sheet(self, sName):
        return self.wb.create_sheet(sName)

    def saveTheWorkBookWithtimestamp(self):
        print("Saving to xlsx file")
        # formats of diffrent date and time to determine the file name:
        my_date = datetime.now().date()
        t = time.localtime()
        current_time = time.strftime("%H_%M_%S", t)
        self._NewFileName = f'D:/ExcelFiles/test_{my_date}_{current_time}.xlsx'
        self.wb.save(self._NewFileName)

    def enterToCell(self, ws, cell, data):
        self.ws[cell] = data

    def getValue(self, cell):
        return self.my_sheet[cell].value

    def change_sheet_title(self, worksheet, new_title):
        worksheet.title = new_title

    def get_headlines(self):
        max_column = openpyxl.utils.get_column_letter(self.my_sheet.max_column)
        allHeaders = {}
        for columnCellObjects in self.my_sheet['A1':max_column + '1']:
            for cellObj in columnCellObjects:
                allHeaders[cellObj.coordinate] = cellObj.value
        return allHeaders


if __name__ == "__main__":
    me = MyExcel()
    print(me.get_sheet_names())
    ws1 = me.my_sheet
    print(me.getValue('B4'))
    me.enterToCell(ws1, 'B4', 'Klemantinot')
    me.change_sheet_title(ws1, 'Ex1')
    ws2 = me.create_new_sheet("MySheet")
    ws2['A4'] = 123
    # scrolling through cells instead of accessing them directly will create them all in memory
    ws2.cell(row=4, column=2, value=10)
    ws2.append([1, 2, 3])
    me.openSheet("MySheet")
    print(me.get_sheet_names())
    me.enterToCell(ws2, 'A2', 'Fruit')
    me.change_sheet_title(ws2, 'Ex2')
    print(me.get_sheet_names())
    me.saveTheWorkBookWithtimestamp()
    print(me.get_headlines())
    print('Done !')